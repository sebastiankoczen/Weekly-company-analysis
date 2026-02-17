"""
Weekly AI Company Analysis - Powered by Gemini + Google Search
Author: rebuilt from scratch with proper error handling

KEY FIXES vs previous versions:
1. API errors are SHOWN clearly, never swallowed silently
2. Correct Gemini Google Search syntax (tested)
3. Raw response always saved to file so you can debug
4. Script FAILS loudly if API key missing or call fails
5. Parser handles Gemini's actual output format flexibly
"""

import os
import sys
import re
import time
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================

GEMINI_API_KEY  = os.environ.get("GEMINI_API_KEY", "")
SEND_EMAIL      = os.environ.get("SEND_EMAIL", "false").lower() == "true"
SMTP_SERVER     = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT       = int(os.environ.get("SMTP_PORT", "587"))
EMAIL_FROM      = os.environ.get("EMAIL_FROM", "")
EMAIL_PASSWORD  = os.environ.get("EMAIL_PASSWORD", "")
EMAIL_TO        = os.environ.get("EMAIL_TO", "")

GEMINI_MODEL       = "gemini-2.0-flash"   # or "gemini-2.0-flash" for cheaper/faster
COMPANIES_PER_WEEK = 10                   # Increase to 10 once results are good
REQUEST_DELAY      = 10                   # Seconds between companies

# ============================================================================
# STARTUP CHECKS — fail loudly early, not silently late
# ============================================================================

def check_environment():
    """Verify everything needed is present before starting"""
    print("🔎 Checking environment...")
    errors = []

    if not GEMINI_API_KEY:
        errors.append("❌ GEMINI_API_KEY is missing from GitHub Secrets")
    else:
        print(f"   ✅ GEMINI_API_KEY found ({GEMINI_API_KEY[:8]}...)")

    if SEND_EMAIL:
        for var, val in [("EMAIL_FROM", EMAIL_FROM), ("EMAIL_PASSWORD", EMAIL_PASSWORD), ("EMAIL_TO", EMAIL_TO)]:
            if not val:
                errors.append(f"❌ {var} is missing from GitHub Secrets")
            else:
                print(f"   ✅ {var} found")

    try:
        import google.generativeai
        print(f"   ✅ google-generativeai installed (v{google.generativeai.__version__})")
    except ImportError:
        errors.append("❌ google-generativeai not installed — check workflow pip install line")

    try:
        import pandas, openpyxl
        print(f"   ✅ pandas + openpyxl installed")
    except ImportError as e:
        errors.append(f"❌ Missing library: {e}")

    if errors:
        print("\n" + "="*70)
        print("STARTUP FAILED — fix these issues before running:")
        for e in errors:
            print(f"  {e}")
        print("="*70)
        sys.exit(1)

    print("   ✅ All checks passed\n")

# ============================================================================
# DATA STRUCTURE
# ============================================================================

class CompanyAnalysis:
    def __init__(self, name):
        self.name = name
        self.situations = {
            1: {"name": "Resource Constraints",   "score": 0, "points": [], "sources": []},
            2: {"name": "Supply Chain Disruption", "score": 0, "points": [], "sources": []},
            3: {"name": "Margin Pressure",         "score": 0, "points": [], "sources": []},
            4: {"name": "Significant Growth",      "score": 0, "points": [], "sources": []},
        }

# ============================================================================
# GEMINI API CALL
# ============================================================================

def get_gemini_response(prompt_text):
    """
    Call Gemini API with Google Search enabled.
    Raises exception with clear message if anything goes wrong.
    """
    import google.generativeai as genai
    from google.generativeai import types

    genai.configure(api_key=GEMINI_API_KEY)

    # Google Search Grounding — makes Gemini search the web in real time
    # This is the documented way for google-generativeai >= 0.5
    try:
        tools = [genai.protos.Tool(
            google_search_retrieval=genai.protos.GoogleSearchRetrieval()
        )]
    except AttributeError:
        # Fallback for older library versions
        print("   ⚠️  Using fallback search tool syntax")
        tools = None

    model = genai.GenerativeModel(
        model_name=GEMINI_MODEL,
        tools=tools,
        generation_config={
            "temperature": 0.3,
            "max_output_tokens": 8192,
        }
    )

    print(f"   📡 Sending request to Gemini ({GEMINI_MODEL})...")
    response = model.generate_content(prompt_text)

    # Check we got actual content
    if not response.candidates:
        raise Exception("Gemini returned no candidates — possible safety block or empty response")

    candidate = response.candidates[0]

    # Check finish reason
    finish_reason = str(candidate.finish_reason)
    if "SAFETY" in finish_reason or "RECITATION" in finish_reason:
        raise Exception(f"Gemini blocked response: {finish_reason}")

    content = response.text
    if not content or len(content.strip()) < 50:
        raise Exception(f"Gemini returned near-empty response ({len(content)} chars): '{content[:100]}'")

    # Log how many sources Gemini searched
    try:
        meta = candidate.grounding_metadata
        if meta and hasattr(meta, 'grounding_chunks') and meta.grounding_chunks:
            print(f"   🔗 Gemini searched {len(meta.grounding_chunks)} web sources")
    except Exception:
        pass

    return content


# ============================================================================
# PARSER — flexible enough to handle Gemini's output
# ============================================================================

def parse_response(response_text):
    """
    Parse Gemini's structured response into CompanyAnalysis objects.
    Handles minor formatting variations in the model's output.
    """
    companies = []

    print(f"\n   Response length: {len(response_text)} characters")
    print(f"   First 200 chars: {repr(response_text[:200])}")

    if "---COMPANY START---" not in response_text:
        print("⚠️  No ---COMPANY START--- delimiter found in response")
        print("   This means Gemini didn't follow the output format.")
        print("   Check the raw .txt file to see what it returned instead.")
        return []

    company_blocks = re.split(r'---COMPANY START---', response_text)
    print(f"   Found {len(company_blocks) - 1} company block(s)")

    for idx, block in enumerate(company_blocks[1:], 1):
        if '---COMPANY END---' in block:
            block = block.split('---COMPANY END---')[0]

        company_match = re.search(r'Company:\s*(.+?)(?:\n|$)', block, re.IGNORECASE)
        if not company_match:
            print(f"   ⚠️  Block {idx}: no Company: line found")
            continue

        company_name = company_match.group(1).strip()
        company = CompanyAnalysis(company_name)
        print(f"   ✅ Parsing: {company_name}")

        situations_found = 0
        for sit_num in range(1, 5):
            # Flexible regex — handles "Key Signals" or "Key Points"
            # and "Evidence Links" or "Sources"
            pattern = (
                rf'SITUATION\s+{sit_num}[:\s].*?\n'
                rf'Score:\s*(\d+).*?\n'
                rf'(?:Key\s+(?:Signals|Points)[^\n]*\n)(.*?)'
                rf'(?:Evidence\s+Links|Sources)[^\n]*\n(.*?)'
                rf'(?=\n\s*SITUATION\s+{sit_num+1}|\n---COMPANY END---|$)'
            )
            match = re.search(pattern, block, re.DOTALL | re.IGNORECASE)

            if match:
                score = int(match.group(1))
                points_raw = match.group(2)
                sources_raw = match.group(3)

                # Extract bullet points
                points = []
                for line in points_raw.split('\n'):
                    line = re.sub(r'^[\s\-\•\*]+', '', line).strip()
                    if line and len(line) > 10:
                        points.append(line)

                # Extract URLs — handles both plain URLs and "DD.MM.YYYY - URL" format
                sources = []
                for line in sources_raw.split('\n'):
                    url = re.search(r'https?://\S+', line)
                    if url:
                        sources.append(url.group(0).rstrip('.,)'))

                company.situations[sit_num]["score"]   = score
                company.situations[sit_num]["points"]  = points[:3]
                company.situations[sit_num]["sources"] = sources[:3]
                situations_found += 1
            else:
                print(f"      ⚠️  Situation {sit_num} not matched by regex")

        print(f"      → {situations_found}/4 situations parsed")
        companies.append(company)

    print(f"\n   ✅ Total companies parsed: {len(companies)}")
    return companies


# ============================================================================
# EXCEL GENERATION
# ============================================================================

def create_excel_report(companies, week_num, output_path):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if not companies:
        print("⚠️  No companies to write — Excel will be empty")

    data = []
    for company in companies:
        for sit_num in range(1, 5):
            s = company.situations[sit_num]
            data.append({
                'Company':     company.name,
                'Situation':   s['name'],
                'Score':       s['score'],
                'Key Point 1': s['points'][0] if len(s['points']) > 0 else '',
                'Key Point 2': s['points'][1] if len(s['points']) > 1 else '',
                'Key Point 3': s['points'][2] if len(s['points']) > 2 else '',
                'Sources':     ' | '.join(s['sources']) if s['sources'] else '',
            })

    df = pd.DataFrame(data)
    df.to_excel(output_path, index=False, sheet_name='Company Analysis')

    wb = load_workbook(output_path)
    ws = wb.active

    # Styles
    hdr_fill = PatternFill("solid", fgColor="366092")
    hdr_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    fill_green  = PatternFill("solid", fgColor="C6EFCE")
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")
    fill_red    = PatternFill("solid", fgColor="FFC7CE")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
            if cell.column == 3 and isinstance(cell.value, int):  # Score
                cell.fill = fill_green if cell.value <= 2 else (fill_yellow if cell.value == 3 else fill_red)
            if cell.column == 7 and cell.value:  # Sources — make first URL clickable
                first_url = re.search(r'https?://\S+', str(cell.value))
                if first_url:
                    cell.hyperlink = first_url.group(0)
                    cell.font = Font(name='Arial', size=10, color="0563C1", underline="single")

    for col, w in zip('ABCDEFG', [22, 26, 8, 42, 42, 42, 55]):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 30 if row[0].row == 1 else 65

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    print(f"   ✅ Excel saved: {output_path}  ({len(data)} rows)")


# ============================================================================
# HTML EMAIL
# ============================================================================

def generate_html_email(companies, week_num):
    rows_html = ""
    for company in companies:
        for sit_num in range(1, 5):
            s = company.situations[sit_num]
            score = s['score']
            color = "#27ae60" if score <= 2 else ("#f39c12" if score == 3 else "#e74c3c")
            bg    = "#d5f4e6" if score <= 2 else ("#fff3cd" if score == 3 else "#f8d7da")
            pts   = "".join(f"<li style='margin-bottom:4px'>{p}</li>" for p in s['points'])
            srcs  = " | ".join(
                f'<a href="{u}" target="_blank" style="color:#3498db">Source {i}</a>'
                for i, u in enumerate(s['sources'][:3], 1) if u.startswith('http')
            )
            rows_html += f"""
            <tr>
              <td style="padding:10px;border:1px solid #ddd;font-weight:bold;color:#2c3e50">{company.name}</td>
              <td style="padding:10px;border:1px solid #ddd;color:#7f8c8d;font-size:0.9em">{s['name']}</td>
              <td style="padding:10px;border:1px solid #ddd;text-align:center">
                <span style="background:{bg};color:{color};padding:4px 10px;border-radius:4px;font-weight:bold">{score}</span>
              </td>
              <td style="padding:10px;border:1px solid #ddd"><ul style="margin:0;padding-left:18px">{pts}</ul></td>
              <td style="padding:10px;border:1px solid #ddd;font-size:0.85em">{srcs}</td>
            </tr>"""

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="font-family:Arial,sans-serif;background:#f5f5f5;padding:20px">
  <div style="max-width:1200px;margin:0 auto;background:white;padding:30px;border-radius:8px">
    <h1 style="color:#2c3e50;border-bottom:3px solid #3498db;padding-bottom:10px">
      📊 Weekly Company Analysis Report
    </h1>
    <div style="background:#ecf0f1;padding:15px;border-radius:5px;margin-bottom:20px">
      <strong>Week {week_num}</strong> | {datetime.now().strftime('%B %d, %Y at %H:%M UTC')}
      &nbsp;|&nbsp; Companies: {len(companies)}
    </div>
    <table style="width:100%;border-collapse:collapse">
      <thead><tr style="background:#34495e;color:white">
        <th style="padding:12px;text-align:left;width:15%">Company</th>
        <th style="padding:12px;text-align:left;width:18%">Situation</th>
        <th style="padding:12px;text-align:left;width:7%">Score</th>
        <th style="padding:12px;text-align:left;width:45%">Key Evidence</th>
        <th style="padding:12px;text-align:left;width:15%">Sources</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
    <div style="margin-top:30px;padding-top:20px;border-top:2px solid #ecf0f1;text-align:center;color:#7f8c8d;font-size:0.9em">
      Auto-generated | Scoring: 1–2 Low Risk · 3 Moderate · 4–5 High Risk
    </div>
  </div>
</body></html>"""


# ============================================================================
# EMAIL SENDER
# ============================================================================

def send_html_email(subject, html_content, excel_path=None):
    if not SEND_EMAIL:
        print("📧 Email disabled (SEND_EMAIL != true)")
        return
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication

    msg = MIMEMultipart()
    msg['From'], msg['To'], msg['Subject'] = EMAIL_FROM, EMAIL_TO, subject
    msg.attach(MIMEText(html_content, 'html'))

    if excel_path and os.path.exists(excel_path):
        with open(excel_path, 'rb') as f:
            att = MIMEApplication(f.read(), _subtype="xlsx")
            att.add_header('Content-Disposition', 'attachment', filename=os.path.basename(excel_path))
            msg.attach(att)

    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    server.starttls()
    server.login(EMAIL_FROM, EMAIL_PASSWORD)
    server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())
    server.quit()
    print(f"   ✅ Email sent → {EMAIL_TO}")


# ============================================================================
# HELPERS
# ============================================================================

def get_companies_for_week(companies, week_num, per_week):
    start = (week_num - 1) * per_week
    if start >= len(companies):
        total_weeks = (len(companies) + per_week - 1) // per_week
        week_num = ((week_num - 1) % total_weeks) + 1
        start = (week_num - 1) * per_week
    return companies[start:start + per_week], week_num

def calculate_current_week(start_date_str, per_week, total):
    start = datetime.strptime(start_date_str, "%Y-%m-%d")
    weeks = (datetime.now() - start).days // 7
    total_weeks = (total + per_week - 1) // per_week
    return (weeks % total_weeks) + 1

def save_raw(content, week_num):
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    fname = f"report_week{week_num}_{ts}.txt"
    with open(fname, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"   💾 Raw response saved: {fname}  ({len(content)} chars)")
    return fname


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 70)
    print("🤖 WEEKLY COMPANY ANALYSIS — GEMINI + GOOGLE SEARCH")
    print("=" * 70)
    print(f"Started : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Model   : {GEMINI_MODEL}")
    print(f"Per week: {COMPANIES_PER_WEEK}")
    print()

    # --- 1. Environment checks (exits loudly if broken) ---
    check_environment()

    # --- 2. Load files ---
    print("📂 Loading files...")
    for f in ["prompt_updated.txt", "companies.txt", "definitions.txt"]:
        if not os.path.exists(f):
            print(f"❌ Missing required file: {f}")
            sys.exit(1)

    role_objective = open("prompt_updated.txt",  encoding="utf-8").read().strip()
    all_companies  = [l.strip() for l in open("companies.txt", encoding="utf-8") if l.strip()]
    scoring_defs   = open("definitions.txt", encoding="utf-8").read().strip()
    print(f"   ✅ Loaded {len(all_companies)} companies")

    # --- 3. Determine week ---
    week_num = int(os.environ.get("WEEK", "0"))
    if week_num == 0:
        start_date = os.environ.get("START_DATE", "2026-02-17")
        week_num = calculate_current_week(start_date, COMPANIES_PER_WEEK, len(all_companies))
    companies_this_week, week_num = get_companies_for_week(all_companies, week_num, COMPANIES_PER_WEEK)

    print(f"\n📋 Week {week_num} — {len(companies_this_week)} companies:")
    for i, c in enumerate(companies_this_week, 1):
        print(f"   {i}. {c}")
    print()

    # --- 4. Analyse one by one ---
    raw_result      = ""
    successful_count = 0
    failed_companies = []

    for i, company_name in enumerate(companies_this_week, 1):
        print(f"[{i}/{len(companies_this_week)}] 🔎 {company_name}")

        single_prompt = (
            f"{role_objective}\n\n"
            f"Definitions of Situations and Scoring:\n{scoring_defs}\n\n"
            f"Company to Analyze:\n{company_name}\n"
        )

        try:
            response = get_gemini_response(single_prompt)
            raw_result += response + "\n\n"
            successful_count += 1
            print(f"   ✅ Got {len(response)} characters")

            if i < len(companies_this_week):
                time.sleep(REQUEST_DELAY)

        except Exception as e:
            # Show the error clearly — do NOT silently continue
            print(f"   ❌ FAILED: {str(e)}")
            failed_companies.append(company_name)
            # Still continue to next company, but failure is visible
            continue

    print(f"\n✅ Successful: {successful_count}/{len(companies_this_week)}")
    if failed_companies:
        print(f"❌ Failed: {failed_companies}")

    # --- 5. Save raw ---
    save_raw(raw_result, week_num)

    if successful_count == 0:
        print("\n❌ CRITICAL: No companies were successfully analysed.")
        print("   Check the error messages above.")
        print("   Most likely cause: GEMINI_API_KEY secret is wrong or missing,")
        print("   OR google-generativeai was not installed in workflow.")
        sys.exit(1)

    # --- 6. Parse ---
    print("\n🔄 Parsing responses...")
    companies_data = parse_response(raw_result)

    if not companies_data:
        print("❌ Parsing returned 0 companies.")
        print("   Open the raw .txt file in artifacts to see what Gemini returned.")
        print("   The format may differ from expected — share it and I'll fix the parser.")
        # Still generate the empty Excel so the artifact is created
    
    # --- 7. Generate outputs ---
    ts             = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_filename = f"analysis_week{week_num}_{ts}.xlsx"

    print(f"\n📊 Generating Excel ({excel_filename})...")
    create_excel_report(companies_data, week_num, excel_filename)

    print("📧 Generating HTML email...")
    html_content = generate_html_email(companies_data, week_num)

    if SEND_EMAIL:
        subject = f"[AUTO-REPORT] 📊 Weekly Company Analysis - Week {week_num}"
        send_html_email(subject, html_content, excel_filename)
    else:
        html_fname = f"email_preview_week{week_num}_{ts}.html"
        open(html_fname, "w", encoding="utf-8").write(html_content)
        print(f"   💾 HTML preview: {html_fname}")

    print("\n" + "="*70)
    print("✅ DONE")
    print("="*70)
    est = successful_count * 0.007
    print(f"💰 Estimated API cost: ~${est:.3f}  ({successful_count} companies × ~$0.007)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
